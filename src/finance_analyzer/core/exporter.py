import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def export_to_excel(df: pd.DataFrame, file_path: str):
    """
    Exports clean, categorized data to Excel.
    Creates distinct visual tables for each category in 'Details' sheet with borders.
    """
    try:
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'
            
        clean_df = df[df['category'].notna() & (df['category'] != '')].copy()
        
        if clean_df.empty:
            return False, "No categorized data to export."

        # Calculate Sums
        cat_sums = clean_df.groupby('category')['amount'].sum().reset_index()
        cat_sums.columns = ['Category', 'Total Amount']
        
        # Sort
        clean_df = clean_df.sort_values(by=['category', 'date'])

        # Create Workbook
        wb = openpyxl.Workbook()
        
        # Styles
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
                             
        # --- Sheet 1: Summary ---
        ws_sum = wb.active
        ws_sum.title = "Summary"
        
        ws_sum.append(['Category', 'Total Amount'])
        for cell in ws_sum[1]:
            cell.font = Font(bold=True)
            cell.border = thin_border
            
        for _, row in cat_sums.iterrows():
            ws_sum.append([row['Category'], row['Total Amount']])
            for cell in ws_sum[ws_sum.max_row]:
                cell.border = thin_border
            
        # --- Sheet 2: Details ---
        ws_det = wb.create_sheet("Details")
        
        headers = ['Date', 'Description', 'Amount', 'Source']
        current_row = 1
        
        categories = clean_df['category'].unique()
        
        for cat in categories:
            cat_data = clean_df[clean_df['category'] == cat]
            cat_total = cat_data['amount'].sum()
            
            # --- Table Start ---
            start_row = current_row
            
            # Category Header (Title)
            ws_det.cell(row=current_row, column=1, value=f"CATEGORY: {cat}")
            ws_det.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
            ws_det.cell(row=current_row, column=1).fill = PatternFill(start_color="1F538D", end_color="1F538D", fill_type="solid")
            ws_det.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            
            # Apply border to merged cell
            for col in range(1, 5):
                ws_det.cell(row=current_row, column=col).border = thin_border
                
            current_row += 1
            
            # Table Header
            for col_idx, h in enumerate(headers, 1):
                cell = ws_det.cell(row=current_row, column=col_idx, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            current_row += 1
            
            # Data Rows
            for _, row in cat_data.iterrows():
                date_str = row['date'].strftime('%Y-%m-%d')
                ws_det.cell(row=current_row, column=1, value=date_str).border = thin_border
                ws_det.cell(row=current_row, column=2, value=row['description']).border = thin_border
                ws_det.cell(row=current_row, column=3, value=row['amount']).border = thin_border
                ws_det.cell(row=current_row, column=4, value=row['source']).border = thin_border
                current_row += 1
                
            # Footer (Sum)
            ws_det.cell(row=current_row, column=2, value="Sum:").border = thin_border
            ws_det.cell(row=current_row, column=2).font = Font(bold=True)
            ws_det.cell(row=current_row, column=2).alignment = Alignment(horizontal='right')
            
            cell_sum = ws_det.cell(row=current_row, column=3, value=cat_total)
            cell_sum.border = thin_border
            cell_sum.font = Font(bold=True)
            cell_sum.number_format = '#,##0.00'
            
            # Empty cells in footer need border? Maybe not to make it look like a "Total" line. 
            # Let's keep it consistent.
            ws_det.cell(row=current_row, column=1).border = thin_border
            ws_det.cell(row=current_row, column=4).border = thin_border
            
            current_row += 3 # spacing between tables

        # Auto-width
        for ws in [ws_sum, ws_det]:
            for column_cells in ws.columns:
                length = max(len(str(cell.value) or "") for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        wb.save(file_path)
        return True, f"Exported to {file_path}"
        
    except Exception as e:
        return False, str(e)
