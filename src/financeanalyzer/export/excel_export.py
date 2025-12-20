"""Excel export service for FinanceAnalyzer."""

from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from ..services.entry_service import EntryService
from ..services.category_service import CategoryService


class ExcelExporter:
    """Service for exporting entries to Excel."""
    
    def __init__(self, profile_id: int):
        """Initialize the exporter.
        
        Args:
            profile_id: The profile ID to export from.
        """
        self.profile_id = profile_id
        
        # Styles
        self.header_font = Font(bold=True, size=12)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_font_white = Font(bold=True, size=12, color="FFFFFF")
        
        self.category_font = Font(bold=True, size=11)
        self.category_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        self.sum_font = Font(bold=True, italic=True)
        self.sum_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.money_positive = Font(color="006400")  # Dark green
        self.money_negative = Font(color="8B0000")  # Dark red
    
    def export(
        self,
        file_path: str | Path,
        export_format: str = "category_tables",
        category_ids: list[int] | None = None,
        include_uncategorized: bool = False,
        sheet_name: str = "Financial Data",
        append_to_existing: bool = False,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> None:
        """Export entries to an Excel file.
        
        Args:
            file_path: Path to save the Excel file.
            export_format: "category_tables" or "all_in_one"
            category_ids: List of category IDs to include, or None for all.
            include_uncategorized: Whether to include uncategorized entries.
            sheet_name: Name for the Excel sheet/tab.
            append_to_existing: If True, append to existing file; else create new.
            start_date: Filter entries on or after this date.
            end_date: Filter entries on or before this date.
        """
        file_path = Path(file_path)
        
        # Get data
        entry_service = EntryService(self.profile_id)
        category_service = CategoryService(self.profile_id)
        
        entries = entry_service.get_all_entries(
            start_date=start_date,
            end_date=end_date
        )
        categories = {c.id: c for c in category_service.get_all_categories()}
        
        entry_service.close()
        category_service.close()
        
        # Filter entries by category
        filtered_entries = []
        for entry in entries:
            # Skip uncategorized if not included
            if entry.category_id is None:
                if include_uncategorized:
                    filtered_entries.append(entry)
                continue
            
            # Filter by category_ids if specified
            if category_ids is None or entry.category_id in category_ids:
                filtered_entries.append(entry)
        
        # Create or open workbook
        if append_to_existing and file_path.exists():
            wb = load_workbook(file_path)
            # Generate unique sheet name if it exists
            base_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base_name}_{counter}"
                counter += 1
            ws = wb.create_sheet(sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]  # Excel limit
        
        # Export based on format
        if export_format == "all_in_one":
            self._export_all_in_one(ws, filtered_entries, categories)
        else:
            self._export_category_tables(ws, filtered_entries, categories)
        
        # Save
        wb.save(file_path)
    
    def _export_category_tables(self, ws, entries: list, categories: dict) -> None:
        """Export entries grouped by category with separate tables."""
        # Group entries by category
        grouped: dict[int | None, list] = {}
        for entry in entries:
            cat_id = entry.category_id
            if cat_id not in grouped:
                grouped[cat_id] = []
            grouped[cat_id].append(entry)
        
        current_row = 1
        grand_total = Decimal("0")
        
        # Sort categories: named categories first (alphabetically), then uncategorized
        sorted_cats = sorted(
            grouped.items(),
            key=lambda x: (x[0] is None, categories.get(x[0], type('', (), {'name': 'ZZZ'})()).name if x[0] else "ZZZ")
        )
        
        for cat_id, cat_entries in sorted_cats:
            # Category header
            if cat_id is None:
                cat_name = "Uncategorized"
            else:
                cat = categories.get(cat_id)
                cat_name = cat.name if cat else f"Unknown ({cat_id})"
            
            ws.cell(row=current_row, column=1, value=f"📁 {cat_name}")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.font = self.category_font
                cell.fill = self.category_fill
                cell.border = self.border
            
            current_row += 1
            
            # Column headers
            headers = ["Date", "Sender/Receiver", "Description", "Source", "Amount"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = self.header_font_white
                cell.fill = self.header_fill
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center')
            
            current_row += 1
            
            # Entries
            cat_total = Decimal("0")
            for entry in sorted(cat_entries, key=lambda e: e.entry_date):
                ws.cell(row=current_row, column=1, value=entry.entry_date.strftime("%d.%m.%Y"))
                ws.cell(row=current_row, column=2, value=getattr(entry, 'sender_receiver', '') or '')
                ws.cell(row=current_row, column=3, value=entry.description[:100])
                ws.cell(row=current_row, column=4, value=entry.source)
                
                amount_cell = ws.cell(row=current_row, column=5, value=float(entry.amount))
                amount_cell.number_format = '#,##0.00 €'
                amount_cell.alignment = Alignment(horizontal='right')
                
                if entry.amount >= 0:
                    amount_cell.font = self.money_positive
                else:
                    amount_cell.font = self.money_negative
                
                for col in range(1, 6):
                    ws.cell(row=current_row, column=col).border = self.border
                
                cat_total += entry.amount
                current_row += 1
            
            # Category subtotal
            ws.cell(row=current_row, column=1, value="Subtotal")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            
            subtotal_cell = ws.cell(row=current_row, column=5, value=float(cat_total))
            subtotal_cell.number_format = '#,##0.00 €'
            subtotal_cell.alignment = Alignment(horizontal='right')
            
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.font = self.sum_font
                cell.fill = self.sum_fill
                cell.border = self.border
            
            grand_total += cat_total
            current_row += 2  # Empty row between categories
        
        # Grand total
        if grouped:
            ws.cell(row=current_row, column=1, value="GRAND TOTAL")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
            
            grand_cell = ws.cell(row=current_row, column=5, value=float(grand_total))
            grand_cell.number_format = '#,##0.00 €'
            grand_cell.alignment = Alignment(horizontal='right')
            
            for col in range(1, 6):
                cell = ws.cell(row=current_row, column=col)
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                cell.border = self.border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def _export_all_in_one(self, ws, entries: list, categories: dict) -> None:
        """Export entries as a pivot table with categories as column headers.
        
        Format: Date | Category1 | Category2 | ... | Total
        Each row is one date-entry combination, amounts in category columns.
        """
        if not entries:
            ws.cell(row=1, column=1, value="No entries to export")
            return
        
        # Get unique categories from entries (sorted alphabetically)
        category_ids_in_entries = set()
        for entry in entries:
            category_ids_in_entries.add(entry.category_id)
        
        # Build ordered category list: named categories first, then uncategorized
        cat_order = []
        for cat_id in sorted(category_ids_in_entries, key=lambda x: (x is None, categories.get(x).name if x and categories.get(x) else "ZZZ")):
            if cat_id is None:
                cat_order.append((None, "Uncategorized"))
            else:
                cat = categories.get(cat_id)
                cat_order.append((cat_id, cat.name if cat else f"Unknown ({cat_id})"))
        
        # Create column mapping: category_id -> column index (1-based, column 1 is Date)
        cat_to_col = {cat_id: i + 2 for i, (cat_id, _) in enumerate(cat_order)}
        total_columns = len(cat_order) + 2  # Date + categories + Total
        
        # Headers
        ws.cell(row=1, column=1, value="Date")
        ws.cell(row=1, column=1).font = self.header_font_white
        ws.cell(row=1, column=1).fill = self.header_fill
        ws.cell(row=1, column=1).border = self.border
        ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
        
        for cat_id, cat_name in cat_order:
            col = cat_to_col[cat_id]
            cell = ws.cell(row=1, column=col, value=cat_name)
            cell.font = self.header_font_white
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Total column header
        total_col = total_columns
        cell = ws.cell(row=1, column=total_col, value="Total")
        cell.font = self.header_font_white
        cell.fill = self.header_fill
        cell.border = self.border
        cell.alignment = Alignment(horizontal='center')
        
        # Sort entries by date
        sorted_entries = sorted(entries, key=lambda e: (e.entry_date, e.description))
        
        # Write entries - each entry gets its own row
        current_row = 2
        column_totals = {cat_id: Decimal("0") for cat_id, _ in cat_order}
        grand_total = Decimal("0")
        
        for entry in sorted_entries:
            # Date
            ws.cell(row=current_row, column=1, value=entry.entry_date.strftime("%d.%m.%Y"))
            ws.cell(row=current_row, column=1).border = self.border
            
            # Amount in category column
            cat_col = cat_to_col.get(entry.category_id, 2)
            amount_cell = ws.cell(row=current_row, column=cat_col, value=float(entry.amount))
            amount_cell.number_format = '#,##0.00 €'
            amount_cell.alignment = Alignment(horizontal='right')
            
            if entry.amount >= 0:
                amount_cell.font = self.money_positive
            else:
                amount_cell.font = self.money_negative
            
            # Apply borders to all cells
            for col in range(1, total_columns + 1):
                ws.cell(row=current_row, column=col).border = self.border
            
            # Row total
            row_total_cell = ws.cell(row=current_row, column=total_col, value=float(entry.amount))
            row_total_cell.number_format = '#,##0.00 €'
            row_total_cell.alignment = Alignment(horizontal='right')
            if entry.amount >= 0:
                row_total_cell.font = self.money_positive
            else:
                row_total_cell.font = self.money_negative
            
            # Track totals
            column_totals[entry.category_id] = column_totals.get(entry.category_id, Decimal("0")) + entry.amount
            grand_total += entry.amount
            
            current_row += 1
        
        # Totals row
        ws.cell(row=current_row, column=1, value="TOTAL")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws.cell(row=current_row, column=1).border = self.border
        
        for cat_id, _ in cat_order:
            col = cat_to_col[cat_id]
            cat_total = column_totals.get(cat_id, Decimal("0"))
            cell = ws.cell(row=current_row, column=col, value=float(cat_total))
            cell.number_format = '#,##0.00 €'
            cell.alignment = Alignment(horizontal='right')
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.border = self.border
        
        # Grand total
        grand_cell = ws.cell(row=current_row, column=total_col, value=float(grand_total))
        grand_cell.number_format = '#,##0.00 €'
        grand_cell.alignment = Alignment(horizontal='right')
        grand_cell.font = Font(bold=True, size=12)
        grand_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        grand_cell.border = self.border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 12
        for i in range(2, total_columns + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15


